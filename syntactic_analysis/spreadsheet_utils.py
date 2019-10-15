# coding: utf-8
from pathlib import Path
import xlrd
from openpyxl import Workbook, load_workbook
import csv
from copy import deepcopy

from .analysis import normalize_raw_tree


def xlsx_to_tsv(filename, out_dir):
    filename, out_dir = Path(filename), Path(out_dir)

    # create and / or empty output folder
    if not out_dir.is_dir():
        out_dir.mkdir(exist_ok=True)
    out_dir = out_dir / filename.stem
    out_dir.mkdir(exist_ok=True)
    for f in out_dir.glob("*.*"):
        f.unlink()

    # write all sheets to temp tsv files
    workbook = xlrd.open_workbook(filename)
    sheets = workbook.sheet_names()
    for s in sheets:
        sheet = workbook.sheet_by_name(s)
        tsv = out_dir / f"{s}.tsv"
        with tsv.open("w", encoding='utf-8-sig') as w:
            writer = csv.writer(w, delimiter="\t")
            for rownum in range(sheet.nrows):
                writer.writerow(sheet.row_values(rownum))


def tsv_to_xlsx(tsv_dir):
    tsv_dir = Path(tsv_dir)
    if not tsv_dir.is_dir():
        raise NotADirectoryError
    if not tsv_dir.glob('*.tsv'):
        raise FileNotFoundError

    workbook = Workbook()
    workbook.remove_sheet(workbook.active)
    for t in sorted(tsv_dir.glob('*.tsv')):
        sheet = workbook.create_sheet(title=t.stem)
        reader = csv.reader(t.open(encoding='utf-8-sig'), delimiter='\t')
        for row in reader:
            sheet.append(row)

    workbook.save(filename=str(tsv_dir) + '.xlsx')


def translate_tsv(tsv):
    en = normalize_raw_tree(deepcopy(tsv), mode='bo_en')
    bo = normalize_raw_tree(deepcopy(tsv), mode='en_bo')
    return en, bo


def translate_trees(filename):
    filename = Path(filename)
    if filename.suffix == '.tsv':
        reader = csv.reader(filename.open(encoding='utf-8-sig'), delimiter='\t')
        en, bo = translate_tsv(list(reader))
        for lang, tree in [('en', en), ('bo', bo)]:
            tsv = filename.parent / f"{filename.stem}_{lang}.tsv"
            with tsv.open("w", encoding='utf-8-sig') as w:
                writer = csv.writer(w, delimiter="\t")
                for row in tree:
                    writer.writerow(row)
    elif filename.suffix == '.xlsx':
        workbook = load_workbook(filename=filename)
        workbook_en = Workbook()
        workbook_en.remove_sheet(workbook_en.active)
        workbook_bo = Workbook()
        workbook_bo.remove_sheet(workbook_bo.active)
        sheets = workbook.sheetnames
        for s in sheets:
            sheet = workbook.get_sheet_by_name(s)
            values = []
            for row in sheet.values:
                row = [r if r else '' for r in row]
                values.append(row)
            en, bo = translate_tsv(values)
            ws_en = workbook_en.create_sheet(s)
            for row in en:
                ws_en.append(row)
            ws_bo = workbook_bo.create_sheet(s)
            for row in bo:
                ws_bo.append(row)

        workbook_en.save(filename=str(filename.parent / filename.stem) + '_en.xlsx')
        workbook_bo.save(filename=str(filename.parent / filename.stem) + '_bo.xlsx')
    else:
        raise NotImplementedError


def translate_tsv_dir(tsv_dir):
    tsv_dir = Path(tsv_dir)
    if not tsv_dir.is_dir():
        raise NotADirectoryError
    if not tsv_dir.glob('*.tsv'):
        raise FileNotFoundError

    en_dir = tsv_dir.parent / (tsv_dir.name + '_en')
    en_dir.mkdir(exist_ok=True)
    bo_dir = tsv_dir.parent / (tsv_dir.name + '_bo')
    bo_dir.mkdir(exist_ok=True)

    for t in tsv_dir.glob('*.tsv'):
        reader = csv.reader(t.open(encoding='utf-8-sig'), delimiter='\t')
        en, bo = translate_tsv(list(reader))
        for lang, tree, dir in [('en', en, en_dir), ('bo', bo, bo_dir)]:
            tsv = dir / f"{t.stem}_{lang}.tsv"
            with tsv.open("w", encoding='utf-8-sig') as w:
                writer = csv.writer(w, delimiter="\t")
                for row in tree:
                    writer.writerow(row)
